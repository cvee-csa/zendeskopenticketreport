#!/usr/bin/env python3
"""
IT Ops Zendesk Tag Report — zendeskreport_opentickets_esc-rarc_bidaily.py

Pulls all open/pending/on-hold tickets from the three IT Ops Zendesk groups,
applies esc/rarc heuristics, and generates a colour-coded Excel report.
The report is uploaded as a GitHub Actions artifact.

Required environment variables:
    ZENDESK_EMAIL   your Zendesk login email
    ZENDESK_TOKEN   Zendesk API token (Admin > Apps & Integrations > API)
"""

import os, re, time, base64, json, html as _html
from datetime import datetime, timezone, timedelta
from urllib.parse import urlencode

import requests
from openpyxl                import Workbook
from openpyxl.styles         import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils          import get_column_letter

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False

# ── Credentials ────────────────────────────────────────────────────────────────
ZENDESK_EMAIL = os.environ["ZENDESK_EMAIL"]
ZENDESK_TOKEN = os.environ["ZENDESK_TOKEN"]

# Google Drive upload (optional — set these secrets to enable)
GDRIVE_SA_JSON  = os.environ.get("GDRIVE_SERVICE_ACCOUNT_JSON")  # full JSON key string
GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID")            # folder or Shared Drive folder ID

# ── Constants ──────────────────────────────────────────────────────────────────
ZENDESK_DOMAIN = "cloudsecurityalliance.zendesk.com"
BASE_ZD        = f"https://{ZENDESK_DOMAIN}/api/v2"
TICKET_URL     = f"https://{ZENDESK_DOMAIN}/agent/tickets/"
PST            = timezone(timedelta(hours=-8))
_now           = datetime.now(PST)
TODAY          = _now.strftime("%Y-%m-%d")
NOW            = _now.strftime("%Y-%m-%d_%I%M") + ("am" if _now.hour < 12 else "pm")
REPORT_PATH    = f"/tmp/IT_Ops_Tag_Report_{NOW}.xlsx"

IT_OPS_GROUPS = {
    7783360594455:  "IT-Operations",
    37981538647191: "IT-Operations-Projects",
    38675924427287: "IT-Operations-Tasks",
}

IT_OPS_ASSIGNEES = {
    19148954105367: "Neeks",
    5720866160535:  "Jacob",
    38942574549655: "Catherine",
}

RYAN_ID = 396710941733
KURT_ID = 396693552053

# ── SLA thresholds (business hours / days) ────────────────────────────────────
# Source: zendesk-parameters.md — "All thresholds apply to business hours only"
SLA_INITIAL_RESPONSE_HRS = 2    # first IT Ops comment within 2 biz hrs of creation
SLA_REQUESTER_WAIT_HRS   = 4    # max biz hours requester waits for IT Ops reply
SLA_NO_UPDATE_HRS        = 8    # 1 biz day = 8 hrs max since any ticket update
SLA_RESOLUTION_DAYS      = 2    # informational: open > 2 biz days flagged

IT_OPS_AGENT_IDS = {19148954105367, 5720866160535, 38942574549655}  # Neeks, Jacob, Catherine

DEADLINES = {
    "security.txt": datetime(2026, 4, 1, tzinfo=timezone.utc),
}

# ── Google Drive upload ────────────────────────────────────────────────────────
def upload_to_gdrive(file_path):
    """
    Upload file_path to Google Drive (works with both My Drive and Shared Drives).
    Requires GDRIVE_SERVICE_ACCOUNT_JSON and GDRIVE_FOLDER_ID env vars.
    Skips silently if either is missing or google-auth libs are not installed.
    """
    if not GDRIVE_AVAILABLE:
        print("  [Drive] google-auth libraries not installed — skipping upload.")
        return
    if not GDRIVE_SA_JSON or not GDRIVE_FOLDER_ID:
        print("  [Drive] GDRIVE_SERVICE_ACCOUNT_JSON or GDRIVE_FOLDER_ID not set — skipping.")
        return

    try:
        sa_json = GDRIVE_SA_JSON.strip()
        if not sa_json:
            print("  [Drive] GDRIVE_SERVICE_ACCOUNT_JSON is blank after stripping whitespace — skipping.")
            return
        creds_info = json.loads(sa_json)

        # Support both service account keys and OAuth user credentials
        if creds_info.get("type") == "service_account":
            creds = service_account.Credentials.from_service_account_info(
                creds_info,
                scopes=["https://www.googleapis.com/auth/drive"],
            )
        else:
            # OAuth user credentials (from get_token.py / InstalledAppFlow)
            from google.oauth2.credentials import Credentials
            creds = Credentials(
                token=creds_info.get("token"),
                refresh_token=creds_info["refresh_token"],
                token_uri=creds_info.get("token_uri", "https://oauth2.googleapis.com/token"),
                client_id=creds_info["client_id"],
                client_secret=creds_info["client_secret"],
                scopes=creds_info.get("scopes"),
            )

        service = build("drive", "v3", credentials=creds)

        file_name = os.path.basename(file_path)
        file_metadata = {"name": file_name, "parents": [GDRIVE_FOLDER_ID]}
        media = MediaFileUpload(
            file_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True,
        )

        # supportsAllDrives=True makes it work for both Shared Drives and My Drive
        uploaded = service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id, name, webViewLink",
            supportsAllDrives=True,
        ).execute()

        print(f"  [Drive] Uploaded: {uploaded['name']}")
        print(f"  [Drive] View at : {uploaded.get('webViewLink', '(no link)')}")

    except json.JSONDecodeError as e:
        print(f"  [Drive] GDRIVE_SERVICE_ACCOUNT_JSON is not valid JSON: {e}")
        print(f"  [Drive] Secret starts with: {repr(GDRIVE_SA_JSON[:80])}")
        print("  [Drive] Check that the secret contains the raw JSON (not base64 or a file path).")
    except Exception as e:
        print(f"  [Drive] Upload failed: {e}")


# ── Zendesk API helpers ─────────────────────────────────────────────────────────
def _zd_headers():
    token = base64.b64encode(
        f"{ZENDESK_EMAIL}/token:{ZENDESK_TOKEN}".encode()
    ).decode()
    return {"Authorization": f"Basic {token}", "Content-Type": "application/json"}




def fetch_tickets():
    """
    Fetch IT Ops open/pending/hold tickets using the cursor-based incremental
    export API (/api/v2/incremental/tickets/cursor.json).

    Unlike the Search API, the incremental endpoint:
      - Has no result-count cap
      - Is not subject to per-agent ticket-visibility restrictions
      - Is the same approach used by the proven export_tickets.py script

    We pull all tickets updated in the last LOOKBACK_DAYS days, then filter
    locally to IT Ops groups + open/pending/hold status.
    """
    LOOKBACK_DAYS = 60
    since = int((datetime.now(timezone.utc) - timedelta(days=LOOKBACK_DAYS)).timestamp())
    print(f"  Using incremental cursor export (last {LOOKBACK_DAYS} days, since={since})")

    url        = f"{BASE_ZD}/incremental/tickets/cursor.json?start_time={since}"
    all_tickets = []
    batch       = 1

    while url:
        r = requests.get(url, headers=_zd_headers(), timeout=60)
        if r.status_code == 429:
            time.sleep(float(r.headers.get("Retry-After", 60)))
            continue
        r.raise_for_status()
        data          = r.json()
        batch_tickets = data.get("tickets", [])
        all_tickets.extend(batch_tickets)
        print(f"  Batch {batch}: {len(batch_tickets)} tickets (running total: {len(all_tickets)})")

        if data.get("end_of_stream", False):
            break
        after_url = data.get("after_url")
        if not after_url or after_url == url:
            break
        url    = after_url
        batch += 1
        time.sleep(0.5)

    # Filter locally to IT Ops groups + actionable statuses
    it_ops_ids     = set(IT_OPS_GROUPS.keys())
    target_statuses = {"open", "pending", "hold"}
    tickets = [
        t for t in all_tickets
        if t.get("group_id") in it_ops_ids
        and t.get("status") in target_statuses
    ]

    print(f"\n  Total from API  : {len(all_tickets)} tickets")
    print(f"  After filtering : {len(tickets)} IT Ops open/pending/hold tickets")
    print(f"\n  --- Sample of fetched tickets (first 5) ---")
    for t in tickets[:5]:
        print(f"    #{t['id']} | group={IT_OPS_GROUPS.get(t.get('group_id'), t.get('group_id'))} | "
              f"status={t.get('status')} | subject={t.get('subject','')[:60]}")
    if not tickets:
        print("  WARNING: 0 IT Ops tickets found after filtering.")
        print(f"  (API returned {len(all_tickets)} total tickets in the window)")
    print(f"  ---\n")

    return tickets


def fetch_comments(ticket_id):
    """Return comment list for a ticket. Retries once on rate-limit."""
    url = f"{BASE_ZD}/tickets/{ticket_id}/comments.json"
    for _ in range(2):
        r = requests.get(url, headers=_zd_headers(), timeout=30)
        if r.status_code == 429:
            time.sleep(float(r.headers.get("Retry-After", 10)))
            continue
        r.raise_for_status()
        return r.json().get("comments", [])
    return []


# ── SLA helpers ────────────────────────────────────────────────────────────────
def _biz_hours_between(start_utc: datetime, end_utc: datetime) -> float:
    """
    Count business hours (Mon–Fri, 09:00–17:00 US/Pacific, UTC-8)
    elapsed between two timezone-aware UTC datetimes.
    Returns a float number of hours (0.0 if end <= start).
    """
    if end_utc <= start_utc:
        return 0.0

    PST_OFFSET = timedelta(hours=-8)   # fixed PST; close enough for SLA flagging
    BIZ_OPEN   = 9
    BIZ_CLOSE  = 17

    # Strip tzinfo and shift to PST so we can do naive date arithmetic
    s = (start_utc + PST_OFFSET).replace(tzinfo=None)
    e = (end_utc   + PST_OFFSET).replace(tzinfo=None)

    total_hours = 0.0
    day = s.replace(hour=0, minute=0, second=0, microsecond=0)
    end_day = e.replace(hour=0, minute=0, second=0, microsecond=0)

    while day <= end_day:
        if day.weekday() < 5:           # Mon=0 … Fri=4
            seg_start = max(s, day.replace(hour=BIZ_OPEN))
            seg_end   = min(e, day.replace(hour=BIZ_CLOSE))
            if seg_end > seg_start:
                total_hours += (seg_end - seg_start).total_seconds() / 3600
        day += timedelta(days=1)

    return total_hours


def _parse_dt(s: str) -> datetime | None:
    """Parse a Zendesk ISO-8601 timestamp string to a timezone-aware datetime."""
    if not s:
        return None
    return datetime.fromisoformat(s.replace("Z", "+00:00"))


def check_sla(ticket: dict, comments: list) -> dict:
    """
    Evaluate SLA compliance for a ticket against the thresholds in
    zendesk-parameters.md.  Returns a dict:
        flags   — list of human-readable flag strings
        display — newline-joined flags, or "✓ Within SLA"
    """
    now        = datetime.now(timezone.utc)
    created_at = _parse_dt(ticket.get("created_at", ""))
    updated_at = _parse_dt(ticket.get("updated_at", ""))
    req_id     = ticket.get("requester_id")

    it_ops_cmts = [c for c in comments if c.get("author_id") in IT_OPS_AGENT_IDS]

    flags = []

    # 1. Initial response: first IT Ops comment within 2 biz hrs of creation
    if created_at:
        if it_ops_cmts:
            first_resp_dt = _parse_dt(it_ops_cmts[0].get("created_at", ""))
            if first_resp_dt:
                hrs = _biz_hours_between(created_at, first_resp_dt)
                if hrs > SLA_INITIAL_RESPONSE_HRS:
                    flags.append(
                        f"⚠ Initial response {hrs:.1f} biz hrs"
                        f" (SLA: {SLA_INITIAL_RESPONSE_HRS} hrs)"
                    )
        else:
            hrs = _biz_hours_between(created_at, now)
            if hrs > SLA_INITIAL_RESPONSE_HRS:
                flags.append(
                    f"🚨 No IT Ops response yet — {hrs:.1f} biz hrs elapsed"
                    f" (SLA: {SLA_INITIAL_RESPONSE_HRS} hrs)"
                )

    # 2. Requester wait: last requester comment unanswered for >4 biz hrs
    if req_id and comments:
        last_req = next((c for c in reversed(comments)
                         if c.get("author_id") == req_id), None)
        if last_req:
            req_dt   = _parse_dt(last_req.get("created_at", ""))
            req_idx  = next((i for i, c in enumerate(comments)
                             if c["id"] == last_req["id"]), -1)
            answered = any(c.get("author_id") in IT_OPS_AGENT_IDS
                           for c in comments[req_idx + 1:])
            if not answered and req_dt:
                hrs = _biz_hours_between(req_dt, now)
                if hrs > SLA_REQUESTER_WAIT_HRS:
                    flags.append(
                        f"🚨 Requester reply unanswered {hrs:.1f} biz hrs"
                        f" (SLA: {SLA_REQUESTER_WAIT_HRS} hrs)"
                    )

    # 3. No update: ticket stale for >1 biz day (8 hrs)
    if updated_at:
        hrs = _biz_hours_between(updated_at, now)
        if hrs > SLA_NO_UPDATE_HRS:
            flags.append(
                f"⚠ No update {hrs:.1f} biz hrs"
                f" (SLA: {SLA_NO_UPDATE_HRS} hrs)"
            )

    # 4. Resolution: informational open-age flag (>2 biz days)
    if created_at:
        open_hrs  = _biz_hours_between(created_at, now)
        open_days = open_hrs / 8
        if open_days > SLA_RESOLUTION_DAYS:
            flags.append(
                f"ℹ Open {open_days:.1f} biz days"
                f" (SLA: {SLA_RESOLUTION_DAYS} days)"
            )

    return {
        "flags":   flags,
        "display": "\n".join(flags) if flags else "✓ Within SLA",
    }


# ── Classification ──────────────────────────────────────────────────────────────
ESC_PATTERNS = [
    r"blocked\s+on",
    r"waiting\s+on\s+(ryan|kurt|dev|r&d|leadership)",
    r"need[s]?\s+(ryan|kurt|dev|leadership|r&d|approval)",
    r"(ryan|kurt)\s+(need[s]?|has\s+to|must|should|is\s+required)",
    r"pending\s+(ryan|kurt|dev|leadership|r&d|approval)",
    r"requires?\s+(ryan|kurt|dev|leadership|approval)",
    r"escalat",
    r"business\s+decision",
    r"leadership\s+decision",
    r"waiting\s+for\s+(a\s+)?(response|decision|approval|review)",
    r"no\s+response\s+(from|since)",
    r"ryan\s+bergsma",
    r"kurt\s+seigfried",
]

RARC_PATTERNS = [
    r"can\s+you\s+confirm",
    r"please\s+confirm",
    r"let\s+me\s+know\s+if",
    r"does\s+this\s+(work|look\s+right|meet)",
    r"is\s+this\s+satisfactory",
    r"can\s+we\s+(go\s+ahead\s+and\s+)?close",
    r"please\s+verify",
    r"good\s+to\s+(go|close)",
    r"everything\s+(look|work|seem)\s+(good|ok|right)",
    r"(ticket|this)\s+can\s+be\s+closed",
    r"let\s+us\s+know\s+when",
    r"confirm.*and\s+(we|i)\s+(will|can|shall)\s+close",
]


# Human-readable descriptions for each pattern
ESC_REASONS = {
    r"blocked\s+on":                                          "Ticket is blocked waiting on someone",
    r"waiting\s+on\s+(ryan|kurt|dev|r&d|leadership)":        "Waiting on Ryan / Kurt / Dev / Leadership",
    r"need[s]?\s+(ryan|kurt|dev|leadership|r&d|approval)":   "Needs input or approval from Ryan / Kurt / Dev / Leadership",
    r"(ryan|kurt)\s+(need[s]?|has\s+to|must|should|is\s+required)": "Action required from Ryan or Kurt",
    r"pending\s+(ryan|kurt|dev|leadership|r&d|approval)":    "Pending action or approval from Ryan / Kurt / Dev / Leadership",
    r"requires?\s+(ryan|kurt|dev|leadership|approval)":      "Requires input or approval",
    r"escalat":                                               "Ticket has been escalated",
    r"business\s+decision":                                  "Awaiting a business decision",
    r"leadership\s+decision":                                "Awaiting a leadership decision",
    r"waiting\s+for\s+(a\s+)?(response|decision|approval|review)": "Waiting for a response, decision, or approval",
    r"no\s+response\s+(from|since)":                         "No response has been received",
    r"ryan\s+bergsma":                                       "Ryan Bergsma is mentioned in the ticket",
    r"kurt\s+seigfried":                                     "Kurt Seigfried is mentioned in the ticket",
}

RARC_REASONS = {
    r"can\s+you\s+confirm":                                  "IT Ops asked the requester to confirm",
    r"please\s+confirm":                                     "IT Ops asked the requester to confirm",
    r"let\s+me\s+know\s+if":                                 "IT Ops is awaiting feedback from the requester",
    r"does\s+this\s+(work|look\s+right|meet)":              "IT Ops asked if the issue has been resolved",
    r"is\s+this\s+satisfactory":                             "IT Ops asked if the resolution is satisfactory",
    r"can\s+we\s+(go\s+ahead\s+and\s+)?close":              "IT Ops asked to close the ticket",
    r"please\s+verify":                                      "IT Ops asked the requester to verify",
    r"good\s+to\s+(go|close)":                               "IT Ops indicated the ticket is ready to close",
    r"everything\s+(look|work|seem)\s+(good|ok|right)":      "IT Ops asked if everything looks good",
    r"(ticket|this)\s+can\s+be\s+closed":                   "IT Ops indicated the ticket can be closed",
    r"let\s+us\s+know\s+when":                               "IT Ops is awaiting confirmation from the requester",
    r"confirm.*and\s+(we|i)\s+(will|can|shall)\s+close":    "IT Ops is waiting to close pending requester confirmation",
}


_EMAIL_RE = re.compile(r"@\S+\.\w{2,}")


def _match_any(patterns, text):
    """Return (pattern, snippet) for the first matching pattern, or (None, None)."""
    # Pre-clean preserving newlines so we can use them as sentence boundaries
    clean = _clean_text(text)
    for p in patterns:
        m = re.search(p, clean, re.IGNORECASE)
        if m:
            before = clean[:m.start()]

            # Walk back to the nearest line / sentence boundary
            best = -1
            for sep in ("\n", ". ", "! ", "? "):
                pos = before.rfind(sep)
                if pos > best:
                    best = pos
            start = (best + 1) if best != -1 else 0

            # If the text between that boundary and the match contains an email
            # address it is almost certainly an email-signature line — skip it
            # and start the snippet from the match itself.
            if _EMAIL_RE.search(clean[start:m.start()]):
                start = m.start()

            # Walk forward to the nearest boundary after the match.
            # Take the minimum (nearest) position across all separator types.
            after_text = clean[m.end():]
            end = len(clean)  # fallback: end of text
            for sep in ("\n", ". ", "! ", "? "):
                pos = after_text.find(sep)
                if pos != -1:
                    candidate = m.end() + pos + len(sep)
                    if candidate < end:
                        end = candidate

            # Collapse newlines to spaces for display — no hard length cap
            snippet = clean[start:end].strip()
            snippet = re.sub(r"\s+", " ", snippet)
            return p, snippet
    return None, None


def classify(ticket, comments):
    """Returns ('esc'|'rarc'|None, reason_str). RARC checked first."""
    status = ticket.get("status", "")

    it_ops_comments = [c for c in comments if c.get("author_id") in IT_OPS_ASSIGNEES]

    if it_ops_comments:
        last_itops     = it_ops_comments[-1]
        last_itops_idx = next(
            (i for i, c in enumerate(comments) if c["id"] == last_itops["id"]), -1
        )
        subsequent        = comments[last_itops_idx + 1:]
        requester_id      = ticket.get("requester_id")
        requester_replied = any(c.get("author_id") == requester_id for c in subsequent)
        matched, snippet  = _match_any(RARC_PATTERNS, last_itops.get("body", ""))
        if matched and not requester_replied:
            description = RARC_REASONS.get(matched, "IT Ops awaiting requester reply")
            return "rarc", f"{description}. No reply from requester yet.\n\u201c{snippet}\u201d"

    all_text = " ".join(
        [ticket.get("subject") or "", ticket.get("description") or ""]
        + [c.get("body", "") for c in comments]
    )
    matched, snippet = _match_any(ESC_PATTERNS, all_text)
    if matched:
        description = ESC_REASONS.get(matched, "Blocked on external actor")
        return "esc", f"{description}.\n\u201c{snippet}\u201d"

    if status == "on-hold":
        return "esc", "Ticket is on-hold — pending external action or decision."

    return None, None


# ── Ryan escalation SLA ─────────────────────────────────────────────────────────
def ryan_escalation(ticket, comments):
    all_text = " ".join(
        [ticket.get("subject", ""), ticket.get("description", "")]
        + [c.get("body", "") for c in comments]
    ).lower()

    if "ryan bergsma" not in all_text and "ryan" not in all_text:
        return ""

    for keyword, deadline in DEADLINES.items():
        if keyword in all_text:
            days_left = (deadline - datetime.now(timezone.utc)).days
            if days_left <= 14:
                return f"Slack Ryan directly — expires {deadline.strftime('%B %-d')}"

    ryan_mentions = [
        c for c in comments
        if re.search(r"ryan", c.get("body", ""), re.IGNORECASE)
    ]
    if not ryan_mentions:
        return "Tag Ryan in ticket"

    last_dt    = max(
        datetime.fromisoformat(c["created_at"].replace("Z", "+00:00"))
        for c in ryan_mentions
    )
    days_since = (datetime.now(timezone.utc) - last_dt).days

    if days_since < 3:  return "Tag Ryan in ticket"
    if days_since < 7:  return "Slack #internal"
    return "Slack Ryan directly"


# ── Last Ryan tag date ─────────────────────────────────────────────────────
def last_ryan_tag_date(comments):
    """
    Return the date (MM/DD/YYYY) of the most recent comment that mentions
    Ryan Bergsma by name, or an empty string if none.
    """
    ryan_cmts = [
        c for c in comments
        if re.search(r"ryan\s+bergsma", c.get("plain_body") or c.get("body") or "", re.IGNORECASE)
    ]
    if not ryan_cmts:
        return ""
    last_dt = max(
        datetime.fromisoformat(c["created_at"].replace("Z", "+00:00"))
        for c in ryan_cmts
    )
    return last_dt.strftime("%m/%d/%Y")


# ── Automated action description ────────────────────────────────────────────────
def _clean_text(text):
    """Strip HTML entities and markdown noise; preserve newlines for boundary detection."""
    text = _html.unescape(text)                              # &nbsp; → space, &amp; → &, etc.
    text = re.sub(r"\*{1,2}([^*]+)\*{1,2}", r"\1", text)  # **bold** / *italic* → plain
    text = re.sub(r"!\[[^\]]*\]\([^)]*\)", "", text)       # remove markdown images
    text = re.sub(r"[ \t]+", " ", text)     # collapse horizontal whitespace only
    text = re.sub(r"\n[ \t]*\n+", "\n", text)  # collapse multiple blank lines to one
    return text.strip()


def _comment_preview(comment, max_chars=150):
    """Return a short, clean excerpt from a comment body."""
    body = (comment.get("plain_body") or comment.get("body") or "")
    body = _clean_text(body)
    if len(body) > max_chars:
        body = body[:max_chars - 1].rstrip() + "…"
    return body


def automated_action(tag, ryan_step, ticket, comments):
    subj     = (ticket.get("subject") or "").lower()
    all_text = " ".join(
        [ticket.get("subject", ""), ticket.get("description", "")]
        + [c.get("body", "") for c in comments]
    ).lower()

    # ── Gather last internal note and last public reply from IT Ops ──────────
    it_ops_cmts   = [c for c in comments if c.get("author_id") in IT_OPS_ASSIGNEES]
    last_internal = next((c for c in reversed(it_ops_cmts) if not c.get("public", True)), None)
    last_public   = next((c for c in reversed(it_ops_cmts) if c.get("public", True)),  None)

    context_lines = []
    if last_internal:
        date    = (last_internal.get("created_at") or "")[:10]
        author  = IT_OPS_ASSIGNEES.get(last_internal.get("author_id"), "IT Ops")
        preview = _comment_preview(last_internal)
        context_lines.append(f"[Internal note — {author}, {date}]\n\"{preview}\"")
    if last_public:
        date    = (last_public.get("created_at") or "")[:10]
        author  = IT_OPS_ASSIGNEES.get(last_public.get("author_id"), "IT Ops")
        preview = _comment_preview(last_public)
        context_lines.append(f"[Public reply — {author}, {date}]\n\"{preview}\"")

    context = "\n".join(context_lines)

    # ── Decide action steps based on tag / escalation state ──────────────────
    if tag == "rarc":
        steps = (
            "1. get_ticket_comments → confirm no requester reply since last IT Ops message.\n"
            "2. If no reply within 3 days: create_ticket_comment (public reply) — send a friendly follow-up asking if the issue is resolved.\n"
            "3. If no reply within 7 days: update_ticket → status = solved, with closing note."
        )
    elif ryan_step == "Tag Ryan in ticket":
        steps = (
            "1. create_ticket_comment (internal note) — @mention Ryan Bergsma with a specific ask referencing the blocked item.\n"
            "2. get_ticket_comments after 48 h — if no acknowledgement: post ticket URL + context to Slack #internal."
        )
    elif ryan_step == "Slack #internal":
        steps = (
            "1. Open Slack → post in #internal: ticket URL + one-line summary of what is blocked and why.\n"
            "2. create_ticket_comment (internal note) — log outreach date and exact message posted."
        )
    elif ryan_step and "directly" in ryan_step and "expires" in ryan_step:
        steps = (
            "1. Send Slack DM to Ryan Bergsma: ticket URL + deadline countdown + specific ask.\n"
            "2. Send Slack DM to Kurt Seigfried with the same context.\n"
            "3. Post to Slack #internal referencing both DMs.\n"
            "4. create_ticket_comment (internal note) — log all outreach dates and messages sent."
        )
    elif ryan_step == "Slack Ryan directly":
        steps = (
            "1. Send Slack DM to Ryan Bergsma: ticket URL + specific ask.\n"
            "2. create_ticket_comment (internal note) — record outreach date and full message sent."
        )
    elif "kurt" in all_text or "seigfried" in all_text:
        steps = (
            "1. create_ticket_comment (internal note) — @mention Kurt Seigfried with specific ask.\n"
            "2. If no acknowledgement within 48 h: post ticket URL to Slack #internal."
        )
    elif any(w in subj for w in ("dev", "code", "workflow", "broken", "fix")):
        steps = (
            "1. get_ticket → retrieve linked Dev ticket; get_ticket_comments for current status.\n"
            "2. create_ticket_comment (public reply) — post status update to requester.\n"
            "3. If Dev ticket has been idle > 5 days: post to Slack #internal with ticket URL."
        )
    else:
        steps = (
            "1. get_ticket_comments → review full thread to identify next actionable step.\n"
            "2. create_ticket_comment (public reply or internal note) — post status update or follow-up."
        )

    if context:
        return f"{context}\n\n{steps}"
    return steps


# ── Spreadsheet builder ─────────────────────────────────────────────────────────
DARK_HEADER = "1F2D3D"
ESC_FILL    = "FFE8E8";  ALT_ESC    = "FFF0F0"
RARC_FILL   = "E8F4E8";  ALT_RARC   = "F0FAF0"
ESC_BADGE   = "C0392B";  RARC_BADGE = "27AE60"
LINK_COLOR  = "1155CC"
AA_BG       = "EEF4FB";  AA_FC      = "1A3A5C"
RYAN_COLORS = {
    "Tag Ryan in ticket":  ("FFF3CD", "5D4037"),
    "Slack #internal":     ("FFD580", "4E342E"),
    "Slack Ryan directly": ("FF8C42", "FFFFFF"),
}
HEADERS = [
    "Tag", "Ticket #", "Group", "Subject", "Reason to Tag",
    "Ticket URL", "Last Updated", "SLA Flags",
    "Last Ryan Tag", "Ryan Escalation", "Automated Actions",
]
WIDTHS = [10, 12, 26, 50, 68, 42, 14, 38, 14, 26, 60]

SLA_BREACH_BG  = "FFF0F0"   # light red — any breach
SLA_URGENT_BG  = "FFD6D6"   # stronger red — 🚨 flags
SLA_OK_BG      = "F0FAF0"   # light green — within SLA


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, bold=False, fc="000000",
          bg=None, wrap=False, align="left", size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fc, size=size)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.border = _border()
    return c


def build_spreadsheet(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tag Recommendations"

    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", start_color=DARK_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 24

    esc_n = rarc_n = ryan_n = 0

    for i, r in enumerate(rows, start=2):
        is_esc  = r["tag"] == "esc"
        even    = i % 2 == 0
        main_bg = (
            (ESC_FILL if is_esc else RARC_FILL) if not even else
            (ALT_ESC  if is_esc else ALT_RARC)
        )
        badge = ESC_BADGE if is_esc else RARC_BADGE
        if is_esc: esc_n  += 1
        else:      rarc_n += 1
        if r["ryan_step"]: ryan_n += 1

        _cell(ws, i, 1, r["tag"],       bold=True, fc=badge,   bg=main_bg, align="center")
        _cell(ws, i, 2, r["ticket_id"], bold=True, fc="333333", bg=main_bg, align="center")
        _cell(ws, i, 3, r["group"],                             bg=main_bg)
        _cell(ws, i, 4, r["subject"],                           bg=main_bg, wrap=True)
        _cell(ws, i, 5, r["reason"],                            bg=main_bg, wrap=True)

        url = f"{TICKET_URL}{r['ticket_id']}"
        lnk = ws.cell(row=i, column=6, value=url)
        lnk.font      = Font(name="Arial", color=LINK_COLOR, underline="single", size=11)
        lnk.alignment = Alignment(horizontal="left", vertical="top")
        lnk.hyperlink = url
        lnk.fill      = PatternFill("solid", start_color=main_bg)
        lnk.border    = _border()

        _cell(ws, i, 7, r["last_updated"], bg=main_bg, align="center")

        # SLA Flags (col 8)
        sla_text  = r.get("sla_display", "✓ Within SLA")
        sla_flags = r.get("sla_flags", [])
        has_urgent = any("🚨" in f for f in sla_flags)
        sla_bg = SLA_URGENT_BG if has_urgent else (SLA_BREACH_BG if sla_flags else SLA_OK_BG)
        sla_fc = "8B0000" if sla_flags else "1B5E20"
        _cell(ws, i, 8, sla_text, fc=sla_fc, bg=sla_bg, wrap=True)

        # Last Ryan Tag date (col 9)
        ryan_date = r.get("last_ryan_tag", "")
        ryan_date_bg = "FFF8E1" if ryan_date else main_bg
        _cell(ws, i, 9, ryan_date, bg=ryan_date_bg, align="center")

        # Ryan Escalation (col 10)
        step = r["ryan_step"]
        if step:
            key      = next((k for k in RYAN_COLORS if step.startswith(k)), None)
            bg2, fc2 = RYAN_COLORS.get(key, ("FF8C42", "FFFFFF"))
            if "expires" in step:
                bg2, fc2 = "E53935", "FFFFFF"
            _cell(ws, i, 10, step, bold=True, fc=fc2, bg=bg2, wrap=True, align="center")
        else:
            _cell(ws, i, 10, "", bg=main_bg, align="center")

        # Automated Actions (col 11)
        _cell(ws, i, 11, r["auto_action"], fc=AA_FC, bg=AA_BG, wrap=True)
        ws.row_dimensions[i].height = 120

    sr = len(rows) + 3
    for col, (val, color) in enumerate([
        (f"Total: {len(rows)} tickets", "1F2D3D"),
        (f"esc: {esc_n}",               ESC_BADGE),
        (f"rarc: {rarc_n}",             RARC_BADGE),
        (f"Ryan bottleneck: {ryan_n}",  "BF360C"),
    ], 1):
        c = ws.cell(row=sr, column=col, value=val)
        c.font = Font(name="Arial", bold=True, size=11, color=color)

    ws.freeze_panes = "A2"
    wb.save(REPORT_PATH)
    print(f"  Spreadsheet saved → {REPORT_PATH}")
    return esc_n, rarc_n, ryan_n


# ── Main ────────────────────────────────────────────────────────────────────────
def main():
    print(f"\n{'='*60}")
    print(f"IT Ops Tag Report — {TODAY}")
    print(f"{'='*60}\n")

    print("[ 1/3 ] Fetching Zendesk tickets...")
    tickets = fetch_tickets()

    print(f"[ 2/3 ] Analysing {len(tickets)} tickets...")
    rows = []
    skipped = 0
    for idx, ticket in enumerate(tickets, 1):
        tid = ticket["id"]
        print(f"  {idx}/{len(tickets)} — #{tid}", end="\r")

        comments    = fetch_comments(tid)
        tag, reason = classify(ticket, comments)
        if not tag:
            skipped += 1
            continue

        group_id = ticket.get("group_id")
        group    = IT_OPS_GROUPS.get(group_id, "IT-Operations")
        updated  = (ticket.get("updated_at") or "")[:10]
        try:
            updated = datetime.strptime(updated, "%Y-%m-%d").strftime("%m/%d/%Y")
        except ValueError:
            pass

        ryan_step     = ryan_escalation(ticket, comments) if tag == "esc" else ""
        last_ryan_tag = last_ryan_tag_date(comments)
        auto          = automated_action(tag, ryan_step, ticket, comments)
        sla           = check_sla(ticket, comments)

        rows.append({
            "tag":           tag,
            "ticket_id":     tid,
            "group":         group,
            "subject":       ticket.get("subject", ""),
            "reason":        reason,
            "last_updated":  updated,
            "sla_flags":     sla["flags"],
            "sla_display":   sla["display"],
            "last_ryan_tag": last_ryan_tag,
            "ryan_step":     ryan_step,
            "auto_action":   auto,
        })
        time.sleep(0.15)

    rows.sort(key=lambda r: (0 if r["tag"] == "esc" else 1, -int(r["ticket_id"])))
    esc_count  = sum(1 for r in rows if r["tag"] == "esc")
    rarc_count = sum(1 for r in rows if r["tag"] == "rarc")
    print(f"\n  {len(rows)} candidates — {esc_count} esc, {rarc_count} rarc "
          f"({skipped} tickets did not match esc/rarc criteria)")

    print("[ 3/3 ] Building spreadsheet...")
    esc_n, rarc_n, ryan_n = build_spreadsheet(rows)
    print(f"  Report: {REPORT_PATH}")

    print("[ + ] Uploading to Google Drive...")
    upload_to_gdrive(REPORT_PATH)

    print(f"\nDone. {len(rows)} tickets reported.\n")


if __name__ == "__main__":
    main()

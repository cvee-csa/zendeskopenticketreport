#!/usr/bin/env python3
"""
IT Ops Zendesk Title Suggester — zendeskreport_opentickets_ticket-title-rename.py

Fetches all open/pending/on-hold tickets from the IT Ops Zendesk groups,
sends each ticket's description to Claude, and generates an Excel report
comparing the current title to Claude's suggested title.

Required environment variables:
    ZENDESK_EMAIL       your Zendesk login email
    ZENDESK_TOKEN       Zendesk API token
    ANTHROPIC_API_KEY   Anthropic API key

Optional environment variables:
    GDRIVE_SERVICE_ACCOUNT_JSON   Google service account or OAuth JSON
    GDRIVE_FOLDER_ID              Target Google Drive folder ID
"""

import os, re, time, base64, json
from datetime import datetime, timezone

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False

# ── Credentials ────────────────────────────────────────────────────────────────
ZENDESK_EMAIL     = os.environ["ZENDESK_EMAIL"]
ZENDESK_TOKEN     = os.environ["ZENDESK_TOKEN"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]

GDRIVE_SA_JSON   = os.environ.get("GDRIVE_SERVICE_ACCOUNT_JSON")
GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID_TITLES")

# ── Constants ──────────────────────────────────────────────────────────────────
ZENDESK_DOMAIN = "cloudsecurityalliance.zendesk.com"
BASE_ZD        = f"https://{ZENDESK_DOMAIN}/api/v2"
TICKET_URL     = f"https://{ZENDESK_DOMAIN}/agent/tickets/"
_now           = datetime.now(timezone.utc)
NOW            = _now.strftime("%Y-%m-%d_%I%M") + ("am" if _now.hour < 12 else "pm")
REPORT_PATH    = f"/tmp/IT_Ops_Title_Suggestions_{NOW}.xlsx"
IT_OPS_GROUPS = {
    7783360594455:  "IT-Operations",
    37981538647191: "IT-Operations-Projects",
    38675924427287: "IT-Operations-Tasks",
}

IT_OPS_ASSIGNEES = {
    19148954105367: "Neeks",
    5720866160535:  "Jacob",
    38942574549655: "Catherine",
}

# ── Zendesk helpers ────────────────────────────────────────────────────────────
def _zd_headers():
    token = base64.b64encode(
        f"{ZENDESK_EMAIL}/token:{ZENDESK_TOKEN}".encode()
    ).decode()
    return {"Authorization": f"Basic {token}", "Content-Type": "application/json"}


def fetch_tickets():
    """Fetch all IT Ops open/pending/hold tickets via the Zendesk search API."""
    group_ids       = list(IT_OPS_GROUPS.keys())
    target_statuses = ["open", "pending", "hold"]
    all_tickets     = []
    seen_ids        = set()

    for status in target_statuses:
        for group_id in group_ids:
            page     = 1
            next_url = None
            while True:
                if next_url:
                    r = requests.get(next_url, headers=_zd_headers(), timeout=60)
                else:
                    r = requests.get(
                        f"{BASE_ZD}/search.json",
                        params={
                            "query":    f"type:ticket status:{status} group_id:{group_id}",
                            "per_page": 100,
                        },
                        headers=_zd_headers(),
                        timeout=60,
                    )
                if r.status_code == 429:
                    time.sleep(float(r.headers.get("Retry-After", 60)))
                    continue
                r.raise_for_status()
                data    = r.json()
                results = data.get("results", [])
                added   = 0
                for t in results:
                    if t["id"] not in seen_ids:
                        seen_ids.add(t["id"])
                        all_tickets.append(t)
                        added += 1
                print(f"  {IT_OPS_GROUPS[group_id]} / {status} — page {page}: "
                      f"{len(results)} results, {added} new")
                next_url = data.get("next_page")
                if not next_url:
                    break
                page += 1
                time.sleep(0.5)

    print(f"\n  Total IT Ops open/pending/hold tickets: {len(all_tickets)}\n")
    return all_tickets


def fetch_ticket_description(ticket_id):
    """Fetch the latest description/comment body for a ticket."""
    url = f"{BASE_ZD}/tickets/{ticket_id}/comments.json?sort_order=asc&per_page=1"
    try:
        r = requests.get(url, headers=_zd_headers(), timeout=30)
        if r.status_code == 429:
            time.sleep(float(r.headers.get("Retry-After", 60)))
            r = requests.get(url, headers=_zd_headers(), timeout=30)
        r.raise_for_status()
        comments = r.json().get("comments", [])
        if comments:
            body = comments[0].get("plain_body") or comments[0].get("body", "")
            # Trim very long descriptions to keep API costs down
            return body[:3000].strip()
    except Exception as e:
        print(f"    Warning: could not fetch description for #{ticket_id}: {e}")
    return ""


# ── Claude title suggestion ────────────────────────────────────────────────────
def suggest_title(client, current_subject, description):
    """Ask Claude to suggest a better ticket title."""
    prompt = f"""You are a Zendesk IT support ticket analyst. Your job is to rewrite ticket titles to be clear, specific, and actionable.

Current ticket title: {current_subject}

Ticket description:
{description if description else "(no description available)"}

Rewrite the ticket title following these rules:
- Must be concise (under 80 characters)
- Must clearly describe the specific issue or request
- Must use plain language (no jargon)
- Must start with an action verb or clear noun phrase (e.g. "Fix...", "Set up...", "Unable to...", "Request: ...", "Create...")
- Must be more specific than the current title
- Never start with "Re:", "Fwd:", "Fw:", or "Notification:"
- Never repeat the current title word-for-word — always rewrite it
- If the description provides more context than the title, use that context

Respond with ONLY the rewritten title — no explanation, no quotes, no punctuation at the end."""

    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=100,
        messages=[{"role": "user", "content": prompt}],
    )
    return message.content[0].text.strip().strip('"').strip("'")


# ── Spreadsheet ────────────────────────────────────────────────────────────────
HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
SAME_FILL   = PatternFill("solid", fgColor="E2EFDA")
BETTER_FILL = PatternFill("solid", fgColor="FFF2CC")
NO_DESC     = PatternFill("solid", fgColor="F2F2F2")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

COLUMNS = [
    ("Ticket #",          10),
    ("Group",             22),
    ("Assignee",          16),
    ("Status",            10),
    ("Current Title",     45),
    ("Suggested Title",   45),
    ("Same?",              8),
    ("Ticket URL",        14),
]


def build_spreadsheet(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Title Suggestions"

    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    changed = same = no_desc = 0

    for row_idx, row in enumerate(rows, start=2):
        alt = row_idx % 2 == 0
        current   = row["current_title"]
        suggested = row["suggested_title"]
        has_desc  = row["has_description"]

        if not has_desc:
            row_fill = NO_DESC
            same_val = "no desc"
            no_desc += 1
        elif suggested.lower() == current.lower():
            row_fill = SAME_FILL
            same_val = "✓"
            same += 1
        else:
            row_fill = BETTER_FILL if not alt else PatternFill("solid", fgColor="FFF9E6")
            same_val = ""
            changed += 1

        values = [
            row["ticket_id"], row["group"], row["assignee"], row["status"],
            current, suggested, same_val, row["url"],
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = row_fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        url_cell = ws.cell(row=row_idx, column=len(COLUMNS))
        url_cell.value     = "Open ticket"
        url_cell.hyperlink = row["url"]
        url_cell.font      = Font(color="1155CC", underline="single")

        ws.row_dimensions[row_idx].height = 40

    summary_row = len(rows) + 2
    summary = ws.cell(
        row=summary_row, column=1,
        value=f"{len(rows)} tickets — {changed} new suggestions / {same} unchanged / {no_desc} no description"
    )
    summary.font = Font(bold=True, italic=True, size=10, color="444444")
    ws.merge_cells(
        start_row=summary_row, start_column=1,
        end_row=summary_row,   end_column=len(COLUMNS)
    )

    wb.save(REPORT_PATH)
    print(f"  Spreadsheet saved → {REPORT_PATH}")
    return changed, same, no_desc


# ── Google Drive upload ────────────────────────────────────────────────────────
def upload_to_gdrive(file_path):
    if not GDRIVE_AVAILABLE:
        print("  [Drive] google-auth libraries not installed — skipping.")
        return
    if not GDRIVE_SA_JSON or not GDRIVE_FOLDER_ID:
        print("  [Drive] GDRIVE_SERVICE_ACCOUNT_JSON or GDRIVE_FOLDER_ID not set — skipping.")
        return
    try:
        creds_info = json.loads(GDRIVE_SA_JSON.strip())
        if creds_info.get("type") == "service_account":
            creds = service_account.Credentials.from_service_account_info(
                creds_info, scopes=["https://www.googleapis.com/auth/drive"]
            )
        else:
            from google.oauth2.credentials import Credentials
            creds = Credentials(
                token=creds_info.get("token"),
                refresh_token=creds_info["refresh_token"],
                token_uri=creds_info.get("token_uri", "https://oauth2.googleapis.com/token"),
                client_id=creds_info["client_id"],
                client_secret=creds_info["client_secret"],
                scopes=creds_info.get("scopes"),
            )
        service   = build("drive", "v3", credentials=creds)
        file_meta = {"name": os.path.basename(file_path), "parents": [GDRIVE_FOLDER_ID]}
        media     = MediaFileUpload(
            file_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True,
        )
        uploaded = service.files().create(
            body=file_meta, media_body=media,
            fields="id, name, webViewLink", supportsAllDrives=True,
        ).execute()
        print(f"  [Drive] Uploaded : {uploaded['name']}")
        print(f"  [Drive] View at  : {uploaded.get('webViewLink', '(no link)')}")
    except Exception as e:
        print(f"  [Drive] Upload failed: {e}")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    if not ANTHROPIC_AVAILABLE:
        raise SystemExit("ERROR: 'anthropic' package not installed. Run: pip install anthropic")

    print(f"\n{'='*60}")
    print(f"IT Ops Title Suggester — {NOW}")
    print(f"{'='*60}\n")

    print("[ 1/3 ] Fetching Zendesk tickets...")
    tickets = fetch_tickets()

    print(f"[ 2/3 ] Generating title suggestions for {len(tickets)} tickets...")
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    rows   = []

    for i, ticket in enumerate(tickets, start=1):
        tid         = ticket["id"]
        subject     = ticket.get("subject") or ticket.get("raw_subject") or "(no title)"
        status      = ticket.get("status", "")
        group       = IT_OPS_GROUPS.get(ticket.get("group_id"), str(ticket.get("group_id")))
        assignee_id = ticket.get("assignee_id")
        assignee    = IT_OPS_ASSIGNEES.get(assignee_id, f"ID:{assignee_id}" if assignee_id else "Unassigned")

        print(f"  {i}/{len(tickets)} — #{tid}  {subject[:60]}")

        # Skip renaming tickets whose title contains square brackets
        has_brackets = "[" in subject or "]" in subject
        if has_brackets:
            print(f"    Skipping — title contains square brackets")

        description = fetch_ticket_description(tid)
        has_desc    = bool(description)

        if has_brackets or not has_desc:
            suggestion = subject
        else:
            try:
                suggestion = suggest_title(client, subject, description)
                # Prefix changed titles with % so they stand out
                if suggestion.lower() != subject.lower():
                    suggestion = f"% {suggestion}"
            except Exception as e:
                err = str(e)
                if "credit balance is too low" in err or "402" in err:
                    print(f"\n  [ERROR] Anthropic account has insufficient credits.")
                    print(f"  [ERROR] Add credits at https://console.anthropic.com/settings/billing")
                    print(f"  [ERROR] Saving partial results ({len(rows)} of {len(tickets)} tickets processed)...")
                    rows.append({
                        "ticket_id":       tid,
                        "group":           group,
                        "assignee":        assignee,
                        "status":          status,
                        "current_title":   subject,
                        "suggested_title": subject,
                        "has_description": has_desc,
                        "url":             f"{TICKET_URL}{tid}",
                    })
                    break
                print(f"    Claude error: {e}")
                suggestion = subject

        rows.append({
            "ticket_id":       tid,
            "group":           group,
            "assignee":        assignee,
            "status":          status,
            "current_title":   subject,
            "suggested_title": suggestion,
            "has_description": has_desc,
            "url":             f"{TICKET_URL}{tid}",
        })

        time.sleep(0.3)

    print(f"\n[ 3/3 ] Building spreadsheet...")
    changed, same, no_desc = build_spreadsheet(rows)
    print(f"  {changed} new suggestions  |  {same} unchanged  |  {no_desc} no description")

    print("\n[ + ] Uploading to Google Drive...")
    upload_to_gdrive(REPORT_PATH)

    print(f"\nDone. Report: {REPORT_PATH}\n")


if __name__ == "__main__":
    main()

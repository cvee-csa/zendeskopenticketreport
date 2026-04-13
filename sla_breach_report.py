#!/usr/bin/env python3
"""
IT Ops SLA Breach Report — sla_breach_report.py

Shows all open IT-Operations tickets, highlights SLA breaches,
tracks when Ryan was last tagged, and recommends the next step
(Slack Ryan the ticket URL, or tag Ryan in the ticket).

Required environment variables:
    ZENDESK_EMAIL   your Zendesk login email
    ZENDESK_TOKEN   Zendesk API token
"""

import os, re, time, base64, json, html as _html
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False

# ── Credentials ──────────────────────────────────────────────────────────────
ZENDESK_EMAIL = os.environ.get("ZENDESK_EMAIL", "")
ZENDESK_TOKEN = os.environ.get("ZENDESK_TOKEN", "")

GDRIVE_SA_JSON   = os.environ.get("GDRIVE_SERVICE_ACCOUNT_JSON")
GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID")

# ── Constants ────────────────────────────────────────────────────────────────
ZENDESK_DOMAIN = "cloudsecurityalliance.zendesk.com"
BASE_ZD        = f"https://{ZENDESK_DOMAIN}/api/v2"
TICKET_URL     = f"https://{ZENDESK_DOMAIN}/agent/tickets/"

_PACIFIC = ZoneInfo("America/Los_Angeles")
_now     = datetime.now(_PACIFIC)
TODAY    = _now.strftime("%Y-%m-%d")
NOW      = _now.strftime("%Y-%m-%d_%I%M") + ("am" if _now.hour < 12 else "pm")
REPORT_PATH = os.environ.get("OUTPUT_FILE", f"/tmp/IT_Ops_SLA_Report_{NOW}.xlsx")

IT_OPS_GROUPS = {
    7783360594455:  "IT-Operations",
    37981538647191: "IT-Operations-Projects",
    38675924427287: "IT-Operations-Tasks",
}

IT_OPS_AGENT_IDS = {19148954105367, 5720866160535, 38942574549655}  # Neeks, Jacob, Catherine
RYAN_ID = 396710941733

# ── SLA thresholds (business hours) ─────────────────────────────────────────
SLA_INITIAL_RESPONSE_HRS = 2   # first IT Ops comment within 2 biz hrs
SLA_REQUESTER_WAIT_HRS   = 4   # max biz hours requester waits unanswered
SLA_NO_UPDATE_HRS        = 8   # 1 biz day = 8 hrs max since any update
SLA_RESOLUTION_DAYS      = 2   # open > 2 biz days flagged


# ── Zendesk API helpers ─────────────────────────────────────────────────────
def _zd_headers():
    token = base64.b64encode(
        f"{ZENDESK_EMAIL}/token:{ZENDESK_TOKEN}".encode()
    ).decode()
    return {"Authorization": f"Basic {token}", "Content-Type": "application/json"}


def fetch_tickets():
    """Fetch IT Ops open/pending/hold tickets via incremental cursor export."""
    LOOKBACK_DAYS = 60
    since = int((datetime.now(timezone.utc) - timedelta(days=LOOKBACK_DAYS)).timestamp())
    print(f"  Fetching tickets (last {LOOKBACK_DAYS} days)...")

    url = f"{BASE_ZD}/incremental/tickets/cursor.json?start_time={since}"
    all_tickets = []
    batch = 1

    while url:
        r = requests.get(url, headers=_zd_headers(), timeout=60)
        if r.status_code == 429:
            time.sleep(float(r.headers.get("Retry-After", 60)))
            continue
        r.raise_for_status()
        data = r.json()
        all_tickets.extend(data.get("tickets", []))
        print(f"  Batch {batch}: {len(data.get('tickets', []))} tickets")
        if data.get("end_of_stream", False):
            break
        after_url = data.get("after_url")
        if not after_url or after_url == url:
            break
        url = after_url
        batch += 1
        time.sleep(0.5)

    it_ops_ids = set(IT_OPS_GROUPS.keys())
    tickets = [
        t for t in all_tickets
        if t.get("group_id") in it_ops_ids
        and t.get("status") in {"open", "pending", "hold"}
    ]
    print(f"  {len(tickets)} IT Ops open/pending/hold tickets (of {len(all_tickets)} total)")
    return tickets


def fetch_comments(ticket_id):
    """Return comment list for a ticket with retry logic."""
    url = f"{BASE_ZD}/tickets/{ticket_id}/comments.json"
    for attempt in range(3):
        r = requests.get(url, headers=_zd_headers(), timeout=30)
        if r.status_code == 429:
            time.sleep(float(r.headers.get("Retry-After", 10)))
            continue
        if r.status_code in (500, 502, 503, 504):
            time.sleep(2 ** attempt)
            continue
        r.raise_for_status()
        return r.json().get("comments", [])
    return []


def fetch_users(user_ids: set) -> dict:
    """Batch-fetch user display names. Returns {user_id: "Name"}."""
    if not user_ids:
        return {}
    users = {}
    ids_list = list(user_ids)
    for i in range(0, len(ids_list), 100):          # ZD allows 100 per call
        batch = ids_list[i:i + 100]
        ids_param = ",".join(str(uid) for uid in batch)
        url = f"{BASE_ZD}/users/show_many.json?ids={ids_param}"
        for attempt in range(3):
            r = requests.get(url, headers=_zd_headers(), timeout=30)
            if r.status_code == 429:
                time.sleep(float(r.headers.get("Retry-After", 10)))
                continue
            if r.ok:
                for u in r.json().get("users", []):
                    users[u["id"]] = u.get("name", f"User #{u['id']}")
                break
    return users


# ── SLA calculation ─────────────────────────────────────────────────────────
def _biz_hours_between(start_utc: datetime, end_utc: datetime) -> float:
    """Count business hours (Mon-Fri, 09:00-17:00 Pacific) between two UTC datetimes."""
    if end_utc <= start_utc:
        return 0.0
    PST_OFFSET = timedelta(hours=-8)
    s = (start_utc + PST_OFFSET).replace(tzinfo=None)
    e = (end_utc + PST_OFFSET).replace(tzinfo=None)
    total = 0.0
    day = s.replace(hour=0, minute=0, second=0, microsecond=0)
    end_day = e.replace(hour=0, minute=0, second=0, microsecond=0)
    while day <= end_day:
        if day.weekday() < 5:
            seg_start = max(s, day.replace(hour=9))
            seg_end = min(e, day.replace(hour=17))
            if seg_end > seg_start:
                total += (seg_end - seg_start).total_seconds() / 3600
        day += timedelta(days=1)
    return total


def _parse_dt(s: str) -> datetime | None:
    if not s:
        return None
    return datetime.fromisoformat(s.replace("Z", "+00:00"))


def check_sla(ticket: dict, comments: list) -> dict:
    """
    Evaluate SLA compliance. Returns:
        breached  — bool, True if any SLA threshold is exceeded
        flags     — list of human-readable flag strings
        severity  — "alert", "warn", or "ok"
        display   — formatted flag string
    """
    now = datetime.now(timezone.utc)
    created_at = _parse_dt(ticket.get("created_at", ""))
    updated_at = _parse_dt(ticket.get("updated_at", ""))
    req_id = ticket.get("requester_id")
    it_ops_cmts = [c for c in comments if c.get("author_id") in IT_OPS_AGENT_IDS]

    flags = []
    severity = "ok"
    no_resp_h = None
    unanswered_h = None
    stale_h = None
    age_days = None

    # 1. Initial response: first IT Ops comment within 2 biz hrs
    if created_at:
        if it_ops_cmts:
            first_resp_dt = _parse_dt(it_ops_cmts[0].get("created_at", ""))
            if first_resp_dt:
                hrs = _biz_hours_between(created_at, first_resp_dt)
                if hrs > SLA_INITIAL_RESPONSE_HRS:
                    unanswered_h = int(hrs)
                    flags.append(f"1st response: {hrs:.0f}h (>{SLA_INITIAL_RESPONSE_HRS}h)")
                    severity = "warn"
        else:
            hrs = _biz_hours_between(created_at, now)
            if hrs > SLA_INITIAL_RESPONSE_HRS:
                no_resp_h = int(hrs)
                unanswered_h = int(hrs)
                flags.append(f"No response: {hrs:.0f}h")
                severity = "alert"

    # 2. Requester wait: unanswered for >4 biz hrs
    if req_id and comments:
        last_req = next((c for c in reversed(comments) if c.get("author_id") == req_id), None)
        if last_req:
            req_dt = _parse_dt(last_req.get("created_at", ""))
            req_idx = next((i for i, c in enumerate(comments) if c["id"] == last_req["id"]), -1)
            answered = any(c.get("author_id") in IT_OPS_AGENT_IDS for c in comments[req_idx + 1:])
            if not answered and req_dt:
                hrs = _biz_hours_between(req_dt, now)
                if hrs > SLA_REQUESTER_WAIT_HRS:
                    unanswered_h = int(hrs)
                    flags.append(f"Unanswered: {hrs:.0f}h (>{SLA_REQUESTER_WAIT_HRS}h)")
                    severity = "alert"

    # 3. No update: stale for >8 biz hrs
    if updated_at:
        hrs = _biz_hours_between(updated_at, now)
        if hrs > SLA_NO_UPDATE_HRS:
            stale_h = int(hrs)
            flags.append(f"Stale: {hrs:.0f}h (>{SLA_NO_UPDATE_HRS}h)")
            if severity == "ok":
                severity = "warn"

    # 4. Resolution: open > 2 biz days (informational)
    if created_at:
        open_hrs = _biz_hours_between(created_at, now)
        open_days = open_hrs / 8
        if open_days > SLA_RESOLUTION_DAYS:
            age_days = int(open_days)
            flags.append(f"Age: {open_days:.0f}d")

    display = " | ".join(flags) if flags else "OK"
    breached = severity in ("warn", "alert")
    return {
        "breached": breached, "flags": flags, "display": display, "severity": severity,
        "no_resp_h": no_resp_h, "unanswered_h": unanswered_h,
        "stale_h": stale_h, "age_days": age_days,
    }


# ── Ryan tracking ───────────────────────────────────────────────────────────
def last_ryan_mention(comments) -> dict:
    """
    Find the most recent comment mentioning Ryan.
    Returns: {"date": "MM/DD/YYYY" or "", "days_ago": int or None, "found": bool}
    """
    ryan_cmts = [
        c for c in comments
        if re.search(r"ryan", c.get("plain_body") or c.get("body") or "", re.IGNORECASE)
    ]
    if not ryan_cmts:
        return {"date": "", "days_ago": None, "found": False}

    last_dt = max(
        datetime.fromisoformat(c["created_at"].replace("Z", "+00:00"))
        for c in ryan_cmts
    )
    days_ago = (datetime.now(timezone.utc) - last_dt).days
    return {"date": last_dt.strftime("%m/%d/%Y"), "days_ago": days_ago, "found": True}


def recommend_next_step(ryan_info: dict, sla: dict) -> str:
    """
    Determine the recommended next step based on Ryan involvement and SLA status.
    Returns one of:
        - "Slack Ryan ticket URL"  (Ryan tagged >3 days ago with no response)
        - "Tag Ryan in ticket"     (Ryan not yet mentioned in the ticket)
        - "Allow time to respond"  (Ryan tagged recently, ≤3 days)
        - ""                       (within SLA, no action needed)
    """
    if not sla["breached"]:
        if ryan_info["found"] and ryan_info["days_ago"] is not None and ryan_info["days_ago"] > 7:
            return "Slack Ryan ticket URL"
        return ""

    # SLA is breached — decide based on Ryan involvement
    if not ryan_info["found"]:
        return "Tag Ryan in ticket"

    days = ryan_info["days_ago"]
    if days is None:
        return "Tag Ryan in ticket"
    if days <= 3:
        return "Allow time to respond"
    return "Slack Ryan ticket URL"


# ── ESC / RARC classification ───────────────────────────────────────────────
ESC_PATTERNS = [
    re.compile(r"blocked\s+on", re.IGNORECASE),
    re.compile(r"waiting\s+on\s+(ryan|kurt|dev|r&d|leadership)", re.IGNORECASE),
    re.compile(r"need[s]?\s+(ryan|kurt|dev|leadership|r&d|approval)", re.IGNORECASE),
    re.compile(r"(ryan|kurt)\s+(need[s]?|has\s+to|must|should|is\s+required)", re.IGNORECASE),
    re.compile(r"pending\s+(ryan|kurt|dev|leadership|r&d|approval)", re.IGNORECASE),
    re.compile(r"requires?\s+(ryan|kurt|dev|leadership|approval)", re.IGNORECASE),
    re.compile(r"escalat", re.IGNORECASE),
    re.compile(r"business\s+decision", re.IGNORECASE),
    re.compile(r"leadership\s+decision", re.IGNORECASE),
    re.compile(r"waiting\s+for\s+(a\s+)?(response|decision|approval|review)", re.IGNORECASE),
    re.compile(r"no\s+response\s+(from|since)", re.IGNORECASE),
    re.compile(r"ryan\s+bergsma", re.IGNORECASE),
    re.compile(r"kurt\s+seigfried", re.IGNORECASE),
]

RARC_PATTERNS = [
    re.compile(r"can\s+you\s+confirm", re.IGNORECASE),
    re.compile(r"please\s+confirm", re.IGNORECASE),
    re.compile(r"let\s+me\s+know\s+if", re.IGNORECASE),
    re.compile(r"does\s+this\s+(work|look\s+right|meet)", re.IGNORECASE),
    re.compile(r"is\s+this\s+satisfactory", re.IGNORECASE),
    re.compile(r"can\s+we\s+(go\s+ahead\s+and\s+)?close", re.IGNORECASE),
    re.compile(r"please\s+verify", re.IGNORECASE),
    re.compile(r"good\s+to\s+(go|close)", re.IGNORECASE),
    re.compile(r"everything\s+(look|work|seem)\s+(good|ok|right)", re.IGNORECASE),
    re.compile(r"(ticket|this)\s+can\s+be\s+closed", re.IGNORECASE),
    re.compile(r"let\s+us\s+know\s+when", re.IGNORECASE),
    re.compile(r"confirm.*and\s+(we|i)\s+(will|can|shall)\s+close", re.IGNORECASE),
]


def classify_esc_rarc(ticket: dict, comments: list) -> str:
    """Classify ticket as 'esc', 'rarc', or '' (neither)."""
    # RARC: check last IT Ops comment for close-ready language
    it_ops_cmts = [c for c in comments if c.get("author_id") in IT_OPS_AGENT_IDS]
    if it_ops_cmts:
        last_itops = it_ops_cmts[-1]
        last_idx = next((i for i, c in enumerate(comments) if c["id"] == last_itops["id"]), -1)
        subsequent = comments[last_idx + 1:]
        req_id = ticket.get("requester_id")
        requester_replied = any(c.get("author_id") == req_id for c in subsequent)
        body = last_itops.get("body") or ""
        if not requester_replied and any(p.search(body) for p in RARC_PATTERNS):
            return "rarc"

    # ESC: check all text for escalation signals
    all_text = " ".join(
        [ticket.get("subject") or "", ticket.get("description") or ""]
        + [c.get("body", "") for c in comments]
    )
    if any(p.search(all_text) for p in ESC_PATTERNS):
        return "esc"

    # On-hold tickets default to ESC
    if ticket.get("status") == "hold":
        return "esc"

    return ""


RYAN_SLACK_HANDLE = "ryanbergsma"


def _summarize_issue(description: str, max_len: int = 120) -> str:
    """Extract a brief plain-text summary from a ticket description.

    Strips HTML tags, collapses whitespace, and truncates to *max_len*
    characters so the Claude prompt gives Ryan context about the original
    request rather than SLA numbers.
    """
    if not description:
        return ""
    # Strip HTML tags
    text = re.sub(r"<[^>]+>", " ", description)
    # Decode common HTML entities
    text = _html.unescape(text)
    # Collapse whitespace, then fix space-before-punctuation
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\s+([.,;:!?])", r"\1", text)
    # Take the first sentence or max_len chars, whichever is shorter
    if len(text) <= max_len:
        return text
    cut = text[:max_len].rsplit(" ", 1)[0]
    return cut + "…" if cut else text[:max_len] + "…"

# Follow-up date: 3 business days from now
_follow_up = _now + timedelta(days=3)
while _follow_up.weekday() >= 5:            # skip weekends
    _follow_up += timedelta(days=1)
FOLLOW_UP_DATE = _follow_up.strftime("%Y-%m-%d")


def build_claude_prompt(ticket_id: int, subject: str, tag: str,
                        next_step: str, issue_summary: str) -> str:
    """Build a copy-paste-ready Claude prompt for ticket automation.

    Each prompt:
      - Opens with 'Use the connected Zendesk tools.'
      - Uses 'First … Then …' for multi-step actions
      - Tags Ryan by @name in Zendesk notes, @handle in Slack DMs
      - Includes a brief summary of the original issue (not SLA numbers)
      - Replaces vague 'check back in 3 days' with a concrete date
    """
    ZD = "Use the connected Zendesk tools."
    parts = []

    # 1. Apply ESC or RARC tag
    if tag == "esc":
        parts.append(f'add the "esc" tag to ticket #{ticket_id}')
    elif tag == "rarc":
        parts.append(f'add the "rarc" tag to ticket #{ticket_id}')

    # Build context blurb: subject + issue summary (skip if near-duplicate)
    context = subject
    if issue_summary:
        subj_norm = subject.lower().strip().rstrip(".")
        summ_norm = issue_summary.lower().strip().rstrip(".")
        if subj_norm not in summ_norm and summ_norm not in subj_norm:
            context = f"{subject}. Issue: {issue_summary}"

    # 2. Next step action
    if next_step == "Tag Ryan in ticket":
        parts.append(
            f'add an internal note to ticket #{ticket_id}: '
            f'"@Ryan Bergsma — {context}. '
            f'Can you review and action this?"'
        )
    elif next_step == "Slack Ryan ticket URL":
        parts.append(
            f'DM @{RYAN_SLACK_HANDLE} in Slack: "Ticket #{ticket_id} — {context}. '
            f'Can you follow up?"'
        )
    # "Allow time to respond" — no Claude action; tag-only prompt is enough.
    # The follow-up date is shown in the Next Step column instead.

    if not parts:
        return ""

    # Sequence with First/Then if multiple actions
    if len(parts) == 1:
        return f"{ZD} {parts[0][0].upper()}{parts[0][1:]}"
    return f"{ZD} First, {parts[0]}. Then {parts[1]}"


# ── Spreadsheet builder — CSA Brand Colors ─────────────────────────────────
CSA_NAVY      = "003366"   # primary — deep navy (logo "CSA" text)
CSA_ORANGE    = "E87722"   # secondary — orange (logo "cloud security alliance")
CSA_BLUE      = "0085CA"   # accent — bright blue (logo bar, links)
CSA_LIGHT     = "E8F1F8"   # light blue tint (backgrounds)
CSA_WHITE     = "FFFFFF"
CSA_DARK_TEXT = "2D3436"   # body text


DARK_HEADER   = CSA_NAVY
BREACH_BG     = "FDE8E8"   # light red for breached rows
BREACH_ALT_BG = "FBDCDC"   # alt row for breached
OK_BG         = CSA_LIGHT  # CSA light blue for OK rows
OK_ALT_BG     = "D6E8F4"   # slightly deeper blue alt
LINK_COLOR    = CSA_BLUE
SUMMARY_BG    = CSA_LIGHT

HEADERS = ["Ticket #", "Subject", "Requester", "SLA", "SLA Detail",
           "Ryan", "Next Step", "Status"]
WIDTHS = [10, 40, 18, 10, 22, 18, 24, 10]

SEVERITY_STYLE = {
    "alert": {"bg": "FBDCDC", "fc": "B71C1C", "label": "BREACH"},
    "warn":  {"bg": "FFF3CD", "fc": "856404", "label": "WARNING"},
    "ok":    {"bg": CSA_LIGHT, "fc": CSA_NAVY, "label": "OK"},
}

STEP_STYLE = {
    "Slack Ryan ticket URL": {"bg": "FBDCDC", "fc": "B71C1C"},
    "Tag Ryan in ticket":    {"bg": "FFF3CD", "fc": "856404"},
    "Allow time to respond": {"bg": CSA_LIGHT, "fc": CSA_NAVY},
    "Follow up in ticket":   {"bg": "FFF3CD", "fc": "856404"},
}


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, bold=False, fc=None,
          bg=None, wrap=False, align="left", size=11):
    if fc is None:
        fc = CSA_DARK_TEXT
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fc, size=size)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.border = _border()
    return c


def write_report(rows: list[dict], output_path: str):
    """Build the styled SLA breach report spreadsheet."""
    wb = Workbook()

    total = len(rows)
    breached = sum(1 for r in rows if r["sla_breached"])
    alerts = sum(1 for r in rows if r["sla_severity"] == "alert")
    warnings = sum(1 for r in rows if r["sla_severity"] == "warn")
    within_sla = total - breached
    ryan_owner = sum(1 for r in rows if r.get("ryan_status") == "Owner")
    ryan_tagged = sum(1 for r in rows if r.get("ryan_status") == "Tagged")
    action_needed = sum(1 for r in rows if r["next_step"])

    sorted_rows = sorted(rows, key=lambda r: (
        0 if r["sla_severity"] == "alert" else 1 if r["sla_severity"] == "warn" else 2,
        -r["days_open"],
    ))

    # ── Summary sheet ────────────────────────────────────────────────
    es = wb.active
    es.title = "Summary"

    es.merge_cells("A1:E1")
    title_cell = es.cell(row=1, column=1, value=f"IT Ops SLA Breach Report — {TODAY}")
    title_cell.font = Font(name="Arial", bold=True, size=16, color=DARK_HEADER)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    es.row_dimensions[1].height = 32

    stats = [
        ("Run Date",          _now.strftime("%Y-%m-%d %H:%M:%S"), CSA_NAVY),
        ("Open Tickets",      total,                               CSA_NAVY),
        ("SLA Breaches",      breached,                            "B71C1C"),
        ("  — Alerts",        alerts,                              "B71C1C"),
        ("  — Warnings",      warnings,                            "856404"),
        ("Within SLA",        within_sla,                          CSA_NAVY),
        ("Ryan — Owner",      ryan_owner,                          CSA_BLUE),
        ("Ryan — Tagged",     ryan_tagged,                         CSA_BLUE),
        ("Action Required",   action_needed,                       "B71C1C"),
    ]
    rn = 3
    for label, val, color in stats:
        es.cell(row=rn, column=1, value=label).font = Font(
            name="Arial", bold=True, size=11, color=CSA_DARK_TEXT)
        v = es.cell(row=rn, column=2, value=val)
        v.font = Font(name="Arial", bold=True, size=13, color=color)
        v.alignment = Alignment(horizontal="left")
        rn += 1

    # SLA thresholds as a one-line footnote
    rn += 1
    footnote = (
        f"SLA Thresholds: 1st Response {SLA_INITIAL_RESPONSE_HRS}h · "
        f"Requester Wait {SLA_REQUESTER_WAIT_HRS}h · "
        f"Stale {SLA_NO_UPDATE_HRS}h (1 day) · "
        f"Resolution Flag {SLA_RESOLUTION_DAYS} days"
    )
    es.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=5)
    es.cell(row=rn, column=1, value=footnote).font = Font(
        name="Arial", size=9, italic=True, color="999999")

    es.column_dimensions["A"].width = 18
    es.column_dimensions["B"].width = 30

    # ── All Open Tickets sheet ────────────────────────────────────────
    ws = wb.create_sheet("All Open Tickets")

    HEADER_ROW = 1
    for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws.cell(row=HEADER_ROW, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", start_color=DARK_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[HEADER_ROW].height = 24

    cur_row = HEADER_ROW + 1
    for r in sorted_rows:
        is_breached = r["sla_breached"]
        even = cur_row % 2 == 0
        bg = (BREACH_BG if not even else BREACH_ALT_BG) if is_breached \
            else (OK_BG if not even else OK_ALT_BG)
        sev = SEVERITY_STYLE[r["sla_severity"]]

        # Col 1: Ticket #
        tid_cell = _cell(ws, cur_row, 1, r["ticket_id"], bold=True,
                         fc=LINK_COLOR, bg=bg, align="center")
        tid_cell.font = Font(name="Arial", bold=True, color=LINK_COLOR,
                             underline="single", size=11)
        tid_cell.hyperlink = r["ticket_url"]

        # Col 2: Subject
        _cell(ws, cur_row, 2, r["subject"], bg=bg, wrap=True)

        # Col 3: Requester
        _cell(ws, cur_row, 3, r.get("requester_name", ""), bg=bg, wrap=True)

        # Col 4: SLA Level
        _cell(ws, cur_row, 4, sev["label"], bold=is_breached,
              fc=sev["fc"], bg=sev["bg"], align="center", size=10)

        # Col 5: SLA Detail (single worst metric)
        detail = r.get("sla_detail", "")
        _cell(ws, cur_row, 5, detail, bold=bool(detail),
              fc=sev["fc"] if detail else CSA_DARK_TEXT,
              bg=sev["bg"] if detail else bg, wrap=True, size=10)

        # Col 6: Ryan status
        ryan_status = r.get("ryan_status", "")
        ryan_days = r["ryan_days_ago"]
        if ryan_status == "Owner":
            ryan_text = "Owner"
            if ryan_days is not None and ryan_days > 0:
                ryan_text += f" (tagged {ryan_days}d)"
            _cell(ws, cur_row, 6, ryan_text, bold=True, fc=CSA_BLUE,
                  bg=bg, align="center")
        elif ryan_status == "Tagged":
            ryan_text = "Tagged"
            if ryan_days is not None and ryan_days > 0:
                ryan_text += f" ({ryan_days}d ago)"
            elif ryan_days == 0:
                ryan_text += " (today)"
            ryan_fc = (CSA_NAVY if ryan_days is not None and ryan_days <= 3
                       else "F57F17" if ryan_days is not None and ryan_days <= 7
                       else "B71C1C")
            _cell(ws, cur_row, 6, ryan_text, bold=True, fc=ryan_fc,
                  bg=bg, align="center")
        else:
            _cell(ws, cur_row, 6, "—", fc="999999", bg=bg, align="center")

        # Col 7: Next Step (show follow-up date for "Allow time" rows)
        step = r["next_step"]
        if step == "Allow time to respond":
            step_text = f"Wait (follow up {FOLLOW_UP_DATE})"
            _cell(ws, cur_row, 7, step_text, bold=True, fc=CSA_NAVY,
                  bg=CSA_LIGHT, wrap=True, size=10)
        elif step:
            ss = STEP_STYLE.get(step, {"bg": bg, "fc": CSA_DARK_TEXT})
            _cell(ws, cur_row, 7, step, bold=True, fc=ss["fc"],
                  bg=ss["bg"], wrap=True)
        else:
            _cell(ws, cur_row, 7, "—", fc=CSA_NAVY, bg=bg, align="center")

        # Col 8: Status
        _cell(ws, cur_row, 8, r["ticket_status"].capitalize(),
              bg=bg, align="center")

        ws.row_dimensions[cur_row].height = 36
        cur_row += 1

    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # ── Claude Prompts sheet (full-width, no truncation) ────────────────
    prompt_rows = [r for r in sorted_rows if r.get("claude_prompt")]
    if prompt_rows:
        ps = wb.create_sheet("Claude Prompts")
        p_headers = ["Ticket #", "Subject", "Next Step", "Claude Prompt"]
        p_widths  = [10, 36, 22, 90]
        for ci, (h, w) in enumerate(zip(p_headers, p_widths), 1):
            c = ps.cell(row=1, column=ci, value=h)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill("solid", start_color=DARK_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()
            ps.column_dimensions[get_column_letter(ci)].width = w
        ps.row_dimensions[1].height = 24

        pr = 2
        for r in prompt_rows:
            tid_cell = _cell(ps, pr, 1, r["ticket_id"], bold=True,
                             fc=LINK_COLOR, align="center")
            tid_cell.font = Font(name="Arial", bold=True, color=LINK_COLOR,
                                 underline="single", size=11)
            tid_cell.hyperlink = r["ticket_url"]

            _cell(ps, pr, 2, r["subject"], wrap=True)
            _cell(ps, pr, 3, r["next_step"] or "—", wrap=True, size=10)
            _cell(ps, pr, 4, r["claude_prompt"], wrap=True, size=10)

            # Dynamic row height based on prompt length
            chars = len(r["claude_prompt"])
            ps.row_dimensions[pr].height = max(36, min(100, chars // 2))
            pr += 1

        ps.freeze_panes = "A2"

    # ── SLA Reference sheet ──────────────────────────────────────────────
    ref = wb.create_sheet("SLA Reference")
    ref.cell(row=1, column=1, value="SLA Thresholds").font = Font(
        name="Arial", bold=True, size=14, color=CSA_NAVY)
    ref.row_dimensions[1].height = 24

    thresholds = [
        ("First Response",  f"{SLA_INITIAL_RESPONSE_HRS} business hours"),
        ("Requester Wait",  f"{SLA_REQUESTER_WAIT_HRS} business hours"),
        ("Stale Ticket",    f"{SLA_NO_UPDATE_HRS} business hours (1 day)"),
        ("Resolution Flag", f"{SLA_RESOLUTION_DAYS} business days"),
    ]
    for i, (label, val) in enumerate(thresholds, 3):
        ref.cell(row=i, column=1, value=label).font = Font(
            name="Arial", bold=True, size=11, color=CSA_DARK_TEXT)
        ref.cell(row=i, column=2, value=val).font = Font(
            name="Arial", size=11, color=CSA_DARK_TEXT)
    ref.column_dimensions["A"].width = 20
    ref.column_dimensions["B"].width = 28

    wb.save(output_path)
    print(f"  Report saved: {output_path} ({len(rows)} tickets)")


# ── Google Drive upload ─────────────────────────────────────────────────────
def upload_to_gdrive(file_path):
    if not GDRIVE_AVAILABLE:
        print("  [Drive] google-auth libraries not installed — skipping.")
        return
    if not GDRIVE_SA_JSON or not GDRIVE_FOLDER_ID:
        print("  [Drive] GDRIVE_SERVICE_ACCOUNT_JSON or GDRIVE_FOLDER_ID not set — skipping.")
        return
    try:
        creds_info = json.loads(GDRIVE_SA_JSON.strip())
        if creds_info.get("type") == "service_account":
            creds = service_account.Credentials.from_service_account_info(
                creds_info, scopes=["https://www.googleapis.com/auth/drive"])
        else:
            from google.oauth2.credentials import Credentials
            creds = Credentials(
                token=creds_info.get("token"),
                refresh_token=creds_info["refresh_token"],
                token_uri=creds_info.get("token_uri", "https://oauth2.googleapis.com/token"),
                client_id=creds_info["client_id"],
                client_secret=creds_info["client_secret"],
            )
        service = build("drive", "v3", credentials=creds)
        name = os.path.basename(file_path)
        metadata = {"name": name, "parents": [GDRIVE_FOLDER_ID]}
        media = MediaFileUpload(file_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        f = service.files().create(body=metadata, media_body=media, fields="id,webViewLink", supportsAllDrives=True).execute()
        print(f"  [Drive] Uploaded: {f.get('webViewLink', f.get('id'))}")
    except Exception as e:
        print(f"  [Drive] Upload failed: {e}")


# ── Main ────────────────────────────────────────────────────────────────────
def main():
    print(f"\n{'='*60}")
    print(f"  IT OPS SLA BREACH REPORT — {TODAY}")
    print(f"{'='*60}\n")

    # 1. Fetch tickets
    print("[1/3] Fetching IT Ops tickets...")
    tickets = fetch_tickets()
    if not tickets:
        print("  No tickets found. Exiting.")
        return

    # 1b. Batch-fetch requester names
    requester_ids = {t.get("requester_id") for t in tickets if t.get("requester_id")}
    print(f"  Fetching {len(requester_ids)} requester names...")
    user_names = fetch_users(requester_ids)

    # 2. Analyze each ticket
    print(f"\n[2/3] Analyzing {len(tickets)} tickets...")
    rows = []
    for i, ticket in enumerate(tickets, 1):
        tid = ticket["id"]
        print(f"  [{i}/{len(tickets)}] #{tid} — {ticket.get('subject', '')[:50]}")

        comments = fetch_comments(tid)
        time.sleep(0.3)  # rate limit courtesy

        # SLA check
        sla = check_sla(ticket, comments)

        # Ryan tracking
        ryan = last_ryan_mention(comments)

        # Next step recommendation
        next_step = recommend_next_step(ryan, sla)

        # ESC/RARC classification
        tag = classify_esc_rarc(ticket, comments)

        # Issue summary for Claude prompts (from ticket description)
        subject = ticket.get("subject", "")
        description = ticket.get("description") or ""
        issue_summary = _summarize_issue(description)

        # Ryan status: Owner / Tagged / —
        ryan_is_owner = ticket.get("assignee_id") == RYAN_ID
        if ryan_is_owner:
            ryan_status = "Owner"
        elif ryan["found"]:
            ryan_status = "Tagged"
        else:
            ryan_status = ""

        # Claude automation prompt
        claude_prompt = build_claude_prompt(tid, subject, tag, next_step, issue_summary)

        # Days open (business hours / 8)
        created_at = _parse_dt(ticket.get("created_at", ""))
        days_open = 0
        if created_at:
            days_open = _biz_hours_between(created_at, datetime.now(timezone.utc)) / 8

        # SLA detail — single worst metric for triage scanning
        if sla["no_resp_h"]:
            sla_detail = f"No response {sla['no_resp_h']}h"
        elif sla["unanswered_h"]:
            sla_detail = f"Unanswered {sla['unanswered_h']}h"
        elif sla["stale_h"]:
            sla_detail = f"Stale {sla['stale_h']}h"
        elif sla["age_days"]:
            sla_detail = f"{sla['age_days']}d old"
        else:
            sla_detail = ""

        # Requester name
        req_id = ticket.get("requester_id")
        requester_name = user_names.get(req_id, "") if req_id else ""

        rows.append({
            "ticket_id":      tid,
            "ticket_url":     f"{TICKET_URL}{tid}",
            "subject":        subject,
            "requester_name": requester_name,
            "ticket_status":  ticket.get("status", ""),
            "days_open":      days_open,
            "sla_breached":   sla["breached"],
            "sla_severity":   sla["severity"],
            "sla_display":    sla["display"],
            "sla_detail":     sla_detail,
            "ryan_found":     ryan["found"],
            "ryan_date":      ryan["date"],
            "ryan_days_ago":  ryan["days_ago"],
            "ryan_status":    ryan_status,
            "next_step":      next_step,
            "tag":            tag,
            "claude_prompt":  claude_prompt,
        })

    # 3. Build report
    print(f"\n[3/3] Building report...")
    write_report(rows, REPORT_PATH)

    # Upload to Google Drive
    upload_to_gdrive(REPORT_PATH)

    # Print summary
    breached = sum(1 for r in rows if r["sla_breached"])
    action = sum(1 for r in rows if r["next_step"])
    print(f"\n{'='*60}")
    print(f"  DONE — {len(rows)} tickets | {breached} SLA breaches | {action} need action")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
